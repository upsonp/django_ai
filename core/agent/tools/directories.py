import csv
import json
from itertools import islice
from pathlib import Path

import logging
from typing import Optional, List, Any

from django.conf import settings
from langchain_core.tools import tool

logger = logging.getLogger(f"ollama.{__name__}")


@tool
def read_excel_file(
        path: str,
        sheet: Optional[int | str] = 0,
        max_rows: int = 50,
        max_cols: int = 20,
        as_dict: bool = True,
        max_bytes: int = 5_000_000,
) -> dict:
    """
    Read/examine an Excel (.xlsx, .xls) or CSV file and return a structured summary + small preview.

    Args:
        path: path to file
        sheet: sheet index (int) or sheet name (str) for Excel files. Ignored for CSV.
               If None, returns info for all sheets (may be larger).
        max_rows: maximum rows to include in preview per sheet
        max_cols: maximum columns to include in preview per sheet
        as_dict: if True, each preview row is a dict mapping column -> value; else a list of values
        max_bytes: maximum allowed file size (bytes) to read/preview

    Returns (dict):
        {
          "status": "ok" | "error",
          "path": "<abs path>",
          "size": <int bytes>,
          "format": "xlsx" | "xls" | "csv" | "unknown",
          "sheets": [
            {
              "name": "<sheet name>",
              "rows": <int total rows if available, else null>,
              "cols": <int total cols if available, else null>,
              "preview": [ {col: value, ...}, ... ]  # up to max_rows rows, truncated to max_cols cols
            },
            ...
          ],
          "requested_sheet": sheet,
          "error": "<message>"  # only present if status == "error"
        }

    Notes:
      - Tries to use pandas if available for convenience; falls back to openpyxl for .xlsx and csv module for .csv.
      - For large files, increase max_bytes or use external processing; defaults aim to keep memory use small.
      - If pandas/openpyxl/xlrd are not installed, behavior falls back and returns an error describing missing deps where needed.
    """
    try:
        p = Path(path)
        logger.debug("Reading file: %s", path)
        if not p.is_file():
            return {"status": "error", "error": f"Not a file: {path}"}

        size = p.stat().st_size
        if size > max_bytes:
            return {"status": "error", "error": f"File too large ({size} bytes) - increase max_bytes to read"}

        suffix = p.suffix.lower()
        fmt = "unknown"
        if suffix == ".csv":
            fmt = "csv"
        elif suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            fmt = "xlsx"
        elif suffix == ".xls":
            fmt = "xls"

        result = {
            "status": "ok",
            "path": str(p.resolve()),
            "size": size,
            "format": fmt,
            "sheets": [],
            "requested_sheet": sheet,
        }

        # Try to use pandas if available (convenient and robust)
        try:
            import pandas as pd  # type: ignore
            has_pandas = True
        except Exception:
            has_pandas = False

        # Helper to trim/format preview rows
        def _build_preview_from_df(df) -> List[Any]:
            # limit columns
            df2 = df.iloc[:, :max_cols] if df.shape[1] > max_cols else df
            # take top max_rows
            df2 = df2.head(max_rows)
            if as_dict:
                # convert NaN -> None for JSON friendliness
                records = df2.where(pd.notna(df2), None).to_dict(orient="records")
                return records
            else:
                return df2.fillna("").astype(str).values.tolist()

        if fmt == "csv":
            # Prefer pandas for csv if available
            if has_pandas:
                # use low_memory to avoid dtype inference issues, only read needed rows
                try:
                    # read header + preview rows
                    df = pd.read_csv(p, nrows=max_rows, dtype=object, low_memory=False)
                    preview = _build_preview_from_df(df)
                    cols = list(df.columns[:max_cols])
                    result["sheets"].append({
                        "name": p.name,
                        "rows": None,  # full row count unknown without a full scan
                        "cols": len(cols),
                        "preview": preview,
                    })
                    return result
                except Exception as e:
                    logger.exception("pandas.read_csv failed, falling back to csv module")
            # fallback to csv module (will read up to max_rows)
            with open(p, "r", encoding="utf-8", errors="replace", newline="") as fh:
                reader = csv.reader(fh)
                preview_rows = []
                header = None
                for i, row in enumerate(reader):
                    if i == 0:
                        header = row[:max_cols]
                        # represent header as columns; preview rows will be dicts if as_dict True
                        if not header:
                            header = [f"col{j}" for j in range(len(row[:max_cols]))]
                    if i > max_rows:
                        break
                    vals = row[:max_cols]
                    if as_dict:
                        # align with header length
                        row_obj = {header[j] if j < len(header) else f"col{j}": vals[j] if j < len(vals) else None
                                   for j in range(min(len(vals), max_cols))}
                        preview_rows.append(row_obj)
                    else:
                        preview_rows.append(vals)
                result["sheets"].append({
                    "name": p.name,
                    "rows": None,
                    "cols": len(header) if header is not None else None,
                    "preview": preview_rows,
                })
                return result

        elif fmt == "xlsx":
            # prefer pandas if available
            if has_pandas:
                try:
                    # If sheet is None -> get all sheet names but only preview each
                    if sheet is None:
                        # get sheet names
                        xls = pd.ExcelFile(p, engine="openpyxl")
                        sheets = xls.sheet_names
                        for sname in sheets:
                            df = pd.read_excel(xls, sheet_name=sname, nrows=max_rows, dtype=object, engine="openpyxl")
                            result["sheets"].append({
                                "name": sname,
                                "rows": None,
                                "cols": df.shape[1],
                                "preview": _build_preview_from_df(df),
                            })
                        return result
                    else:
                        df = pd.read_excel(p, sheet_name=sheet, nrows=max_rows, dtype=object, engine="openpyxl")
                        result["sheets"].append({
                            "name": df.columns.name if getattr(df, "columns", None) is not None else str(sheet),
                            "rows": None,
                            "cols": df.shape[1],
                            "preview": _build_preview_from_df(df),
                        })
                        return result
                except Exception as e:
                    logger.exception("pandas.read_excel failed; will try openpyxl fallback")

            # pandas not available or failed -> try openpyxl directly
            try:
                from openpyxl import load_workbook  # type: ignore
            except Exception as e:
                return {"status": "error",
                        "error": "pandas not installed and openpyxl not available to read .xlsx files"}

            wb = load_workbook(filename=str(p), read_only=True, data_only=True)
            names = wb.sheetnames
            target_sheets = names if sheet is None else ([names[sheet]] if isinstance(sheet, int) else [str(sheet)])
            for sname in target_sheets:
                if sname not in names:
                    # skip unknown sheet names
                    continue
                ws = wb[sname]
                preview_rows = []
                max_col_seen = 0
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= max_rows:
                        break
                    # row is a tuple of values
                    vals = list(row[:max_cols])
                    max_col_seen = max(max_col_seen, len(vals))
                    if as_dict:
                        # build header from first row if available; otherwise use col0..colN
                        # here we'll treat the first row as header only if i==0; otherwise use default col names
                        if i == 0:
                            header = [str(x) if x is not None else f"col{j}" for j, x in enumerate(vals)]
                            # do not append header row as data
                            continue
                        # Map using header if present
                        if 'header' in locals():
                            row_obj = {header[j] if j < len(header) else f"col{j}": vals[j] if j < len(vals) else None
                                       for j in range(min(len(vals), max_cols))}
                        else:
                            row_obj = {f"col{j}": vals[j] if j < len(vals) else None for j in
                                       range(min(len(vals), max_cols))}
                        preview_rows.append(row_obj)
                    else:
                        preview_rows.append(vals)
                # estimate cols/rows where possible
                result["sheets"].append({
                    "name": sname,
                    "rows": None,
                    "cols": max_col_seen if max_col_seen > 0 else None,
                    "preview": preview_rows,
                })
            return result

        elif fmt == "xls":
            # .xls older format: try pandas if available, else try xlrd
            if has_pandas:
                try:
                    df = pd.read_excel(p, sheet_name=sheet if sheet is not None else 0, nrows=max_rows)
                    result["sheets"].append({
                        "name": str(sheet or 0),
                        "rows": None,
                        "cols": df.shape[1],
                        "preview": _build_preview_from_df(df),
                    })
                    return result
                except Exception:
                    logger.exception("pandas.read_excel for .xls failed")
            # try xlrd fallback
            try:
                import xlrd  # type: ignore
            except Exception:
                return {"status": "error", "error": "pandas not installed and xlrd not available to read .xls files"}

            wb = xlrd.open_workbook(str(p))
            sheet_names = wb.sheet_names()
            target = sheet if sheet is not None else 0
            if isinstance(target, int):
                sname = sheet_names[target] if target < len(sheet_names) else None
            else:
                sname = target if target in sheet_names else None
            if sname is None:
                return {"status": "error", "error": "Requested sheet not found"}
            sh = wb.sheet_by_name(sname)
            preview_rows = []
            cols = min(max_cols, sh.ncols)
            # optionally treat first row as header when building dicts
            header = [str(sh.cell_value(0, c)) if sh.nrows > 0 else f"col{c}" for c in range(cols)]
            for r in range(1, min(sh.nrows, max_rows + 1)):
                vals = [sh.cell_value(r, c) for c in range(cols)]
                if as_dict:
                    row_obj = {header[c]: vals[c] for c in range(len(vals))}
                    preview_rows.append(row_obj)
                else:
                    preview_rows.append(vals)
            result["sheets"].append({
                "name": sname,
                "rows": sh.nrows,
                "cols": sh.ncols,
                "preview": preview_rows,
            })
            return result

        else:
            return {"status": "error", "error": f"Unsupported file extension: {suffix}"}

    except Exception as e:
        logger.exception("read_excel_file error")
        return {"status": "error", "error": str(e)}


@tool
def ls_recursive_list_paginated(path: str, page: int = 0, page_size: int = 100) -> dict:
    """
    Return page `page` (0-indexed) of files with `page_size`.
    This implementation enumerates results deterministically using sorted ordering.
    """
    logger.debug("MEDIA_IN dir: %s", settings.MEDIA_IN)
    root = Path(settings.MEDIA_IN) if path == '/' else Path(settings.MEDIA_IN, path)
    logger.debug("listing path: %s", root)
    if not root.is_dir():
        return {"status": "error", "error": f"Not a directory: {path}"}

    # Collect paths deterministically (sorted) but you can optimize by scanning and counting
    all_files = sorted(str(p) for p in root.rglob("*") if p.is_file())

    start = page * page_size
    end = start + page_size
    page_files = all_files[start:end]

    files = []
    for fp in page_files:
        p = Path(fp)
        stat = p.stat()
        files.append({
            "path": fp,
            "name": p.name,
            "is_file": True,
            "size": stat.st_size,
            "mtime": stat.st_mtime,
        })

    next_page: Optional[int] = page + 1 if end < len(all_files) else None

    return {
        "status": "ok",
        "root": str(root),
        "count": len(all_files),
        "files": files,
        "next_page": next_page,
    }


def _ls_recursive_list_structured(path: str) -> dict:
    """
    Returns a structured dict describing files under path.
    """
    logger.debug("MEDIA_IN dir: %s", settings.MEDIA_IN)
    if not path or path == '/':
        root = Path(settings.MEDIA_IN)
    else:
        media_root = Path(settings.MEDIA_IN).resolve()
        rel_path = str(Path(path.lstrip('/')))
        root = (media_root / rel_path).resolve()

    logger.debug("listing path: %s", root)
    if not root.is_dir():
        return {"status": "error", "error": f"Not a directory: {root}"}

    files = []
    for p in root.rglob("*"):
        logger.debug("listing path: %s", p)
        if p.is_file():
            stat = p.stat()
            files.append({
                "path": str(p),
                "name": p.name,
                "is_file": True,
                "size": stat.st_size,
                "mtime": stat.st_mtime,
            })

    return {
        "status": "ok",
        "root": str(root),
        "count": len(files),
        "files": files,
        "next_page": None,
    }


@tool
def ls_recursive_list(path: str) -> dict:
    """Recursively explores directories from a provided root directory and returns a structured dictionary of
    directories and files
    Args:
        path (str): path to index from
    Returns:
        dict: List of directories found at path or a dictionary with an error message
    """
    try:
        result = _ls_recursive_list_structured(path)
        try:
            serialized = json.dumps(result)
            logger.debug("ls_recursive_list returning result type=%s serialized_bytes=%d file_count=%d",
                         type(result).__name__, len(serialized), result.get("count", -1))
        except Exception:
            logger.debug("ls_recursive_list returning result type=%s (could not JSON-serialize for debug)",
                         type(result).__name__)
        return result
    except Exception as e:
        logger.exception("ls_recursive_list error while listing path: %s", path)
        return {"status": "error", "error": str(e)}


@tool
def ls_files_structured(path: str, absolute: bool = False) -> dict:
    """
    Lists only files within a directory and returns a structured response for an A.I agent

    Returns:
        {
          "status": "ok" | "error",
          "root": "<root path>",
          "count": <int>,
          "files": [ {"path": "...", "name": "...", "size": <int>, "mtime": <float>} ],
          "error": "<message>"  # only present if status == "error"
        }
    """
    logger.info(f"Structured listing at path: {path} (absolute={absolute})")
    try:
        root = Path(path)
        if not root.is_dir():
            return {"status": "error", "error": f"Not a directory: {path}"}

        files = []
        subdir_count = 0
        for p in sorted(root.iterdir()):
            if p.is_file():
                stat = p.stat()
                files.append({
                    "path": str(p) if absolute else str(p.relative_to(root)),
                    "name": p.name,
                    "size": stat.st_size,
                    "mtime": stat.st_mtime,
                })
            elif p.is_dir():
                subdir_count += 1

        return {
            "status": "ok",
            "root": str(root),
            "file count": len(files),
            "files": files,
        }
    except Exception as e:
        logger.error("listing path: %s", str(e))
        return {"status": "error", "error": str(e)}


@tool
def list_missions() -> dict:
    """List directories in the mission folder, each directory is considered a mission on its own
    Returns:
        dict: dictionary of elements or error messages
    """
    try:
        root = Path(settings.MEDIA_IN).resolve()
        missions = [
            {"name": str(mission), "path": str(mission.resolve())}
            for mission in root.iterdir() if mission.is_dir()
        ]
        return {
            "status": "ok",
            "missions": missions,
        }
    except Exception as e:
        logger.exception("Error listing missions %s", str(e))
        return {"status": "error", "error": str(e)}


@tool
def list_mission_directories(mission: str) -> dict:
    """ Reports ONLY directories under a mission folder.
    Args:
        mission (str): mission name
    Returns:
        dict: dictionary of elements or error messages
    """
    try:
        logger.debug("MEDIA_IN dir: %s", settings.MEDIA_IN)
        media_root = Path(settings.MEDIA_IN).resolve()
        root = (media_root / mission).resolve()

        directories = []
        for p in root.rglob("*"):
            if p.is_dir():
                directories.append({
                    "path": str(p),
                })

        result = {
            "status": "ok",
            "mission": mission,
            "subdirectories": directories,
        }
        return result
    except Exception as e:
        logger.exception("listing path: %s", str(e))
        return {"status": "error", "error": str(e)}


@tool
def list_mission_files(mission: str, page: int = 0) -> str:
    """
    Returns a structured dict describing folders and files under a mission directory.
    Args:
        mission (str): name of the mission
        page (int, optional): page number
    Returns:
        dict: dictionary of elements or error messages
    """
    try:
        page_size: int = 10

        logger.debug("MEDIA_IN dir: %s", settings.MEDIA_IN)
        media_root = Path(settings.MEDIA_IN).resolve()
        root = (media_root / mission).resolve()

        logger.debug("listing path: %s", root)
        if not str(root).startswith(str(media_root)):
            return {"status": "error", "error": "Invalid mission path"}

        if not root.is_dir():
            return {"status": "error", "error": f"Not a mission: {mission}"}

        def _file_iter():
            for p in root.rglob("*"):
                if p.is_file():
                    stat = p.stat()
                    yield {
                        "path": str(p),
                        "name": p.name,
                        "is_file": True,
                        "size": stat.st_size,
                        "mtime": stat.st_mtime,
                    }

        start = page * page_size
        stop = start + page_size
        files_page = list(islice(_file_iter(), start, stop))

        # Indicate whether there is a next page (try to peek one more item)
        has_next = False
        if len(files_page) == page_size:
            # try to see if there is another file after the page (cheap one-item scan)
            next_item_iter = islice(_file_iter(), stop, stop + 1)
            has_next = bool(list(next_item_iter))

            result = {
                "status": "ok",
                "mission": mission,
                "page": page,
                "page_size": page_size,
                "files": files_page,
                "has_next": has_next,
            }
        return result
    except Exception as e:
        logger.exception("listing path: %s", str(e))
        return {"status": "error", "error": str(e)}


@tool
def open_text_file(path: str, max_chars: int = 200_000, tail: bool = False,
                              encoding: str | None = None) -> dict:
    """
    Read a plain-text file such as (.log, .csv, .dat) with safety limits and return a structured dict.

    Args:
        path: file path
        max_chars: max characters to return (truncates if larger)
        tail: if True, return the last max_chars of the file instead of the head
        encoding: optional encoding to use; if None, tries utf-8 with errors='replace'

    Returns:
        {
          "status": "ok" | "error",
          "path": "<abs path>",
          "size": <int bytes>,
          "encoding": "<used encoding>",
          "truncated": <bool>,
          "content": "<string>",
          "error": "<message>"  # only if status == "error"
        }
    """
    try:
        p = Path(path)
        logger.debug("reading file: %s", path)
        if not p.is_file():
            return {"status": "error", "error": f"Not a file: {path}"}

        size = p.stat().st_size
        used_encoding = encoding or "utf-8"

        # Read in binary and decode safely to avoid crashes on bad encodings.
        with open(p, "rb") as f:
            if not tail:
                raw = f.read(max_chars * 4 + 1)  # read a bit more to be safe for multibyte chars
            else:
                # seek to near end
                to_read = min(size, max_chars * 4 + 1)
                f.seek(max(0, size - to_read))
                raw = f.read(to_read)

        try:
            text = raw.decode(used_encoding, errors="strict")
        except Exception:
            # fall back to replacing undecodable bytes
            text = raw.decode(used_encoding, errors="replace")

        # If we read more bytes than max_chars (account for multibyte decode), truncate by characters
        if len(text) > max_chars:
            if tail:
                content = text[-max_chars:]
            else:
                content = text[:max_chars]
            truncated = True
        else:
            content = text
            truncated = (size > len(content))

        return {
            "status": "ok",
            "path": str(p.resolve()),
            "size": size,
            "encoding": used_encoding,
            "truncated": truncated,
            "content": content,
        }
    except Exception as e:
        logger.exception("read_text_file_structured error")
        return {"status": "error", "error": str(e)}

